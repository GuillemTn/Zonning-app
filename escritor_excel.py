"""
Módulo escritor_excel.py

Responsabilidad única: Generar archivo Excel de salida.
- Toma diccionario de asignaciones {(empleado, hora): zona}
- Genera DataFrame con formato: Empleado X Horas (07:00-22:00)
- Escribe a Zonning_Hoy.xlsx, hoja "Cuadrante"
"""

import pandas as pd
import os
from copy import copy
import openpyxl
import unicodedata
import tempfile
import shutil
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _extraer_grupos_por_color(ruta_horario='Horario semana.xlsx'):
    """
    Extrae los grupos de empleados basándose en el color de fondo de la celda.
    Grupo 1: Termina en EBF0DE
    Grupo 2: Termina en E3DFEB
    """
    grupos = {}
    if not os.path.exists(ruta_horario):
        return grupos
    try:
        wb = openpyxl.load_workbook(ruta_horario, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Buscamos en las primeras columnas (usualmente la 1 o 2 contiene los nombres)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
                for cell in row:
                    val = str(cell.value).strip() if cell.value else ''
                    if val and cell.fill and cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                        color = str(cell.fill.fgColor.rgb).upper()
                        if color.endswith('EBF0DE'):
                            grupos[val] = 1
                        elif color.endswith('E3DFEB'):
                            grupos[val] = 2
        return grupos
    except Exception as e:
        print(f"[Advertencia] Error al leer colores de {ruta_horario}: {e}")
        return grupos


def _leer_tabla_mapeo_estilos(ruta_plantilla='Plantilla_Visual.xlsx', hoja='Config_Formatos'):
    """
    Lee la hoja de configuración `Config_Formatos` en la plantilla y construye
    un diccionario: ID_Sistema -> {'nombre': Nombre_Visual, 'style_cell': cell}

    Espera una tabla con columnas (nombres insensibles a mayúsculas):
    - ID_Sistema
    - Nombre_Visual
    - Estilo_Visual (celda cuyo estilo será copiado)

    Retorna:
        dict, openpyxl.Workbook
    """
    if not os.path.exists(ruta_plantilla):
        raise FileNotFoundError(f"Plantilla no encontrada: {ruta_plantilla}")

    wb = openpyxl.load_workbook(ruta_plantilla)
    if hoja not in wb.sheetnames:
        raise ValueError(f"Hoja '{hoja}' no encontrada en {ruta_plantilla}")

    ws = wb[hoja]

    # Localizar fila de cabecera (primera fila que contiene 'ID' y 'Nombre')
    header_row = 1
    headers = {}
    for r in range(1, 6):
        row_vals = [str(ws.cell(row=r, column=c).value).strip() if ws.cell(row=r, column=c).value is not None else '' for c in range(1, ws.max_column + 1)]
        low = [v.lower() for v in row_vals]
        if any('id' in v for v in low) and any('nombre' in v for v in low):
            header_row = r
            for c, v in enumerate(row_vals, start=1):
                if v:
                    headers[v.strip().lower()] = c
            break

    # Normalizar posibles nombres de columna; por defecto asumimos A=ID, B=Nombre
    col_id = None
    col_nombre = None
    col_estilo = None
    for k, c in headers.items():
        if 'id' in k:
            col_id = c
        if 'nombre' in k:
            col_nombre = c
        if 'est' in k:  # Estilo_Visual
            col_estilo = c

    # Fallbacks: si no detectamos cabeceras, usamos columna A/B por defecto
    if col_id is None:
        col_id = 1
    if col_nombre is None:
        col_nombre = 2
    # Si no hay columna Estilo_Visual, usaremos la columna Nombre_Visual como fuente de estilo
    if col_estilo is None:
        col_estilo = col_nombre

    mapping = {}
    for r in range(header_row + 1, ws.max_row + 1):
        id_val = ws.cell(row=r, column=col_id).value
        if id_val is None:
            continue
        id_key = str(id_val).strip()
        nombre_val = ws.cell(row=r, column=col_nombre).value
        nombre_str = str(nombre_val).strip() if nombre_val is not None else id_key
        estilo_cell = ws.cell(row=r, column=col_estilo)

        # Construir objeto estilo copiando componentes con copy() para evitar referencias
        try:
            style_obj = {
                'font': copy(estilo_cell.font),
                'fill': copy(estilo_cell.fill),
                'border': copy(estilo_cell.border),
                'alignment': copy(estilo_cell.alignment),
                'number_format': estilo_cell.number_format
            }
        except Exception:
            # En caso de celdas vacías o estilos no aplicables, usar None para cada componente
            style_obj = {
                'font': None,
                'fill': None,
                'border': None,
                'alignment': None,
                'number_format': None
            }

        mapping[id_key] = {
            'nombre': nombre_str,
            'style': style_obj
        }

    return mapping, wb



def generar_cuadrante(empleados_hoy, cuadrante_dict):
    """
    Convierte diccionario {(empleado, hora): zona} a DataFrame.
    
    Estructura:
    - Columna 0: Empleado
    - Columnas 1-16: Horas 07:00-22:00
    
    Retorna:
        DataFrame
    """
    nombre_col = empleados_hoy.columns[0]

    # Filtrar empleados que trabajan hoy (Estado != 'OFF') si la columna existe
    if 'Estado' in empleados_hoy.columns:
        mask = empleados_hoy['Estado'].astype(str).str.upper() != 'OFF'
        empleados = empleados_hoy[mask][nombre_col].tolist()
        entradas = empleados_hoy[mask]['Hora de entrada'].tolist()
    else:
        empleados = empleados_hoy[nombre_col].tolist()
        entradas = empleados_hoy['Hora de entrada'].tolist()
    
    # Crear matriz: filas = empleados, columnas = horas
    horas = [f"{h}-{h+1}" for h in range(7, 22)]
    
    data = {'Empleado': empleados}
    data['HORA ENTR'] = entradas
    for hora in horas:
        data[hora] = [''] * len(empleados)
    
    df = pd.DataFrame(data)
    
    # Rellenar zonas asignadas
    for (emp_name, hora_int), zona in cuadrante_dict.items():
        hora_str = f"{hora_int}-{hora_int+1}"
        
        # Buscar fila del empleado
        idx = df[df['Empleado'] == emp_name].index
        if len(idx) > 0:
            df.at[idx[0], hora_str] = zona
    
    return df


def generar_cuadrante_buscador(empleados_hoy, cuadrante_dict):
    """Versión segura con búsqueda que maneja índices correctamente."""
    nombre_col = empleados_hoy.columns[0]
    # Mapeo: empleado_name -> índice en la lista
    if 'Estado' in empleados_hoy.columns:
        mask = empleados_hoy['Estado'].astype(str).str.upper() != 'OFF'
        empleados_lista = empleados_hoy[mask][nombre_col].tolist()
        entradas_lista = empleados_hoy[mask]['Hora de entrada'].tolist()
        # Capturar índice original si existe
        indices_lista = empleados_hoy[mask]['original_idx'].tolist() if 'original_idx' in empleados_hoy.columns else []
        salidas_lista = empleados_hoy[mask]['Hora_salida'].tolist() if 'Hora_salida' in empleados_hoy.columns else []
    else:
        empleados_lista = empleados_hoy[nombre_col].tolist()
        entradas_lista = empleados_hoy['Hora de entrada'].tolist()
        indices_lista = empleados_hoy['original_idx'].tolist() if 'original_idx' in empleados_hoy.columns else []
        salidas_lista = empleados_hoy['Hora_salida'].tolist() if 'Hora_salida' in empleados_hoy.columns else []
            
    idx_map = {emp: i for i, emp in enumerate(empleados_lista)}
    
    # Preparar estructura de datos
    horas = [f"{h}-{h+1}" for h in range(7, 22)]
    datos_grid = [[''] * len(horas) for _ in range(len(empleados_lista))]
    
    # Rellenar con asignaciones
    for (emp_name, hora_int), zona in cuadrante_dict.items():
        if emp_name in idx_map:
            emp_idx = idx_map[emp_name]
            hora_idx = hora_int - 7  # Columna 0 = hora 7, columna 15 = hora 22
            if 0 <= hora_idx < len(horas):
                datos_grid[emp_idx][hora_idx] = zona
    
    # Crear DataFrame
    df = pd.DataFrame(datos_grid, columns=horas)
    df.insert(0, 'Empleado', empleados_lista)
    df.insert(1, 'HORA ENTR', entradas_lista)
    if salidas_lista:
        df['HORA SALIDA'] = salidas_lista
    if indices_lista:
        df['original_idx'] = indices_lista
    
    return df


import math

def _parsear_hora_a_entero(hora_str):
    """Convierte 'HH:MM' a float (7.0, 7.5 para 7:30). Retorna None si falla."""
    if pd.isna(hora_str) or str(hora_str).strip() == '':
        return None
    try:
        hora_str = str(hora_str).strip()
        if ':' in hora_str:
            parts = hora_str.split(':')
            hour = int(parts[0])
            minute = int(parts[1])
            return hour + minute / 60.0
        else:
            return float(hora_str)
    except:
        return None

def calcular_descansos(empleados_df):
    """
    Calcula los descansos para empleados que trabajan 6h o más.
    - REGLA ESPECIAL: Tsotsoriia, Georgii siempre a las 19:30 si es posible.
    - Prioriza descanso a las 3.5h del inicio.
    - No más de 2 personas por slot de 30 min.
    - Retorna un dict: {empleado: "HH:MM"}
    """
    nombre_col = empleados_df.columns[0]
    descansos = {}
    slots_ocupados = {} # Formato {13.5: 2} para el slot de 13:30

    # 1. Filtrar empleados elegibles
    elegibles = []
    for _, emp_row in empleados_df.iterrows():
        h_entrada_f = _parsear_hora_a_entero(emp_row.get('Hora de entrada'))
        h_salida_f = _parsear_hora_a_entero(emp_row.get('Hora_salida'))

        if h_entrada_f is not None and h_salida_f is not None:
            duracion = h_salida_f - h_entrada_f
            if duracion >= 6:
                elegibles.append({
                    'nombre': emp_row[nombre_col],
                    'entrada': h_entrada_f,
                    'salida': h_salida_f
                })

    # 2. Manejar casos especiales (Tsotsoriia, Georgii) y separar al resto
    georgii_case = None
    otros_elegibles = []
    for emp in elegibles:
        if 'tsotsoriia, georgii' in emp['nombre'].lower():
            georgii_case = emp
        else:
            otros_elegibles.append(emp)
    
    # Ordenar el resto por hora de entrada para que la asignación sea consistente
    otros_elegibles.sort(key=lambda x: x['entrada'])

    # Asignar a Georgii primero si existe y es posible
    if georgii_case:
        # Comprobar si las 19:30 está dentro de su turno y hay hueco
        if georgii_case['entrada'] <= 19.5 and georgii_case['salida'] > 19.5 and slots_ocupados.get(19.5, 0) < 2:
            descansos[georgii_case['nombre']] = "19:30"
            slots_ocupados[19.5] = slots_ocupados.get(19.5, 0) + 1

    # 3. Asignar slots para el resto de empleados
    for emp in otros_elegibles:
        # Hora ideal: 30min despues de la mitad de turno -> [entrada] + (horas trabajadas+30min)/2
        duracion = emp['salida'] - emp['entrada']
        hora_ideal_f = emp['entrada'] + (duracion + 0.5) / 2
        slot_ideal = round(hora_ideal_f * 2) / 2.0

        slot_asignado = None
        
        # Buscar slot desde el ideal hacia adelante
        slot_actual = slot_ideal
        while slot_actual < emp['salida'] - 0.5: # El descanso debe terminar antes de la última media hora
            if slots_ocupados.get(slot_actual, 0) < 2:
                slot_asignado = slot_actual
                break
            slot_actual += 0.5 # Probar siguiente slot de 30 mins

        if slot_asignado is not None:
            slots_ocupados[slot_asignado] = slots_ocupados.get(slot_asignado, 0) + 1
            
            # Convertir float a formato "HH:MM"
            hora = int(slot_asignado)
            minutos = "30" if slot_asignado % 1 != 0 else "00"
            descansos[emp['nombre']] = f"{hora:02d}:{minutos}"

    return descansos


def escribir_excel(df, ruta_salida='Zonning_Hoy.xlsx'):
    """
    Escribe DataFrame a Excel en hoja "Cuadrante".
    
    Parámetros:
        df: DataFrame a escribir
        ruta_salida: Ruta del archivo (default: Zonning_Hoy.xlsx)
    
    Retorna:
        bool: True si exitoso, False si error
    """
    grupos_map = _extraer_grupos_por_color('Horario semana.xlsx')

    try:
        # Aquí usamos la plantilla para aplicar estilos
        mapping, wb = _leer_tabla_mapeo_estilos()

        # Construir un índice normalizado (sin acentos, mayúsculas) para
        # permitir detectar variantes como 'CAMIÓN' vs 'CAMION' o diferencias
        # de mayúsculas/minúsculas en el ID de sistema.
        def _norm_key(s):
            if s is None:
                return ''
            # Normalizar unicode (quitar acentos), pasar a ASCII y mayúsculas
            nk = unicodedata.normalize('NFKD', str(s))
            nk = nk.encode('ASCII', 'ignore').decode('ASCII')
            return nk.strip().upper()

        mapping_norm = { _norm_key(k): v for k, v in mapping.items() }

        # Obtener (o crear) hoja 'Cuadrante'
        if 'Cuadrante' in wb.sheetnames:
            ws = wb['Cuadrante']
            # limpiar contenido existente (mantener estilos de plantilla si existen filas/cols)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.value = None
                    # No limpiar estilos aquí; sólo se aplicarán donde haya datos.
        else:
            ws = wb.create_sheet('Cuadrante')

        # Configurar Zoom al 65%
        ws.sheet_view.zoomScale = 65
        ws.row_dimensions[1].height = 40

        # Escribir encabezados: Empleado + 07:00..22:00
        headers = ['Empleado', 'HORA ENTR'] + [f"{h}-{h+1}" for h in range(7, 22)]
        
        # Definir estilos reutilizables
        header_fill = PatternFill(start_color="1E4A80", end_color="1E4A80", fill_type="solid")
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        entry_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        group1_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
        group2_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        special_hour_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
        special_hours = ['7-8', '8-9', '9-10', '21-22']
        header_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        emp_font = Font(name='Calibri', size=18)
        entry_font = Font(name='Calibri', size=16, bold=True)
        body_font = Font(name='Calibri', size=11)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        emp_alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
        thin_border = Border(left=Side(style='thin', color='808080'), 
                             right=Side(style='thin', color='808080'), 
                             top=Side(style='thin', color='808080'), 
                             bottom=Side(style='thin', color='808080'))
        emp_border = Border(left=Side(style='thin', color='808080'),
                            right=Side(style='thin', color='808080'),
                            top=Side(style='thin', color='9BC2E6'),
                            bottom=Side(style='thin', color='9BC2E6'))
        
        # ESCRITURA DE ENCABEZADOS CON ESTILOS
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            # Aplicar negrita a todos los encabezados
            cell.font = header_font
            cell.border = thin_border
            # Aplicar color gris claro y alineación centrada a la fila de horas
            cell.fill = header_fill
            cell.alignment = center_alignment

        # Escribir filas
        num_filas = len(df)
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            ws.row_dimensions[r_idx].height = 23.4
            # Empleado con fondo gris claro y alineación centrada
            emp_cell = ws.cell(row=r_idx, column=1, value=row['Empleado'])
            emp_cell.font = emp_font
            emp_cell.alignment = emp_alignment
            emp_cell.border = emp_border

            # Aplicar color según color de fondo en Horario semana.xlsx
            emp_name_str = str(row['Empleado']).strip()
            grupo = grupos_map.get(emp_name_str, 0)
            if grupo == 1:
                emp_cell.fill = group1_fill
            elif grupo == 2:
                emp_cell.fill = group2_fill
            else:
                emp_cell.fill = gray_fill
            
            # HORA ENTR
            entr_cell = ws.cell(row=r_idx, column=2, value=row['HORA ENTR'])
            entr_cell.font = entry_font
            entr_cell.fill = entry_fill
            entr_cell.alignment = center_alignment
            entr_cell.number_format = 'HH:mm'
            entr_cell.border = thin_border

            for c_idx, hora in enumerate(headers[2:], start=3):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                # Aplicar fondo gris suave a columnas especiales (7, 8, 9, 21)
                if hora in special_hours:
                    cell.fill = special_hour_fill
                
                val = row.get(hora, '')
                if val is None or str(val).strip() == '':
                    # Dejar en blanco pero aplicar alineación centrada
                    cell.value = ''
                    cell.alignment = center_alignment
                    continue
                # No escribir la palabra OFF; dejar en blanco cuando el empleado no está activo
                if str(val).strip().upper() == 'OFF':
                    cell.value = ''
                    cell.alignment = center_alignment
                    continue

                # Si existe mapeo por ID, usar nombre visual y estilo
                key = str(val).strip()
                style_entry = mapping.get(key)
                if style_entry is None:
                    # Intentar búsqueda normalizada (quitar acentos / diferencias de caso)
                    style_entry = mapping_norm.get(_norm_key(key))
                if style_entry:
                    write_val = style_entry['nombre']
                    cell.value = write_val
                    s = style_entry['style']
                    try:
                        f = copy(s['font'])
                        if f:
                            f.name = 'Calibri'
                        cell.font = f
                        cell.fill = copy(s['fill'])
                        # Usar borde de plantilla si existe, si no el fino
                        cell.border = copy(s['border']) if s['border'] and s['border'].left.style else thin_border
                        # Mantener alineación centrada incluso con estilos de mapeo
                        cell.alignment = center_alignment
                        cell.number_format = s.get('number_format', cell.number_format)
                    except Exception:
                        # En caso de error al aplicar estilo, aplicar al menos alineación
                        cell.alignment = center_alignment
                else:
                    # No hay mapeo: escribir el valor tal cual con alineación centrada
                    cell.value = val
                    cell.font = body_font
                    cell.alignment = center_alignment

        # Ajustar ancho de columnas para mejor visualización
        ws.column_dimensions['A'].width = 48  # Columna de empleados
        ws.column_dimensions['B'].width = 12   # Columna de entrada
        for c in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 12  # Columnas de horas

        # Guardar en nueva ruta
        wb.save(ruta_salida)
        print(f"[OK] Archivo generado: {ruta_salida}")
        return True

    except Exception as e:
        print(f"[ERROR] No se pudo escribir {ruta_salida}: {e}")
        return False


def generar_salida_semanal(empleados_por_dia, cuadrantes_por_dia, ruta_salida, ruta_plantilla):
    """
    Genera un archivo Excel consolidado con una pestaña por día, usando estilos de Plantilla_Visual.xlsx
    
    Parámetros:
        empleados_por_dia: Dict {día: DataFrame empleados}
        cuadrantes_por_dia: Dict {día: dict cuadrante}
        ruta_salida: Ruta del archivo final
    
    Retorna:
        bool: True si exitoso, False si error
    """
    if not cuadrantes_por_dia:
        print("[ERROR] No hay cuadrantes para escribir")
        return False
    
    try:
        # Cargar mapeo de estilos y workbook de la plantilla
        mapping, wb = _leer_tabla_mapeo_estilos(ruta_plantilla)
        
        # Limpiar hojas existentes (excepto Config_Formatos)
        for hoja in list(wb.sheetnames):
            if hoja.lower() != 'config_formatos':
                del wb[hoja]
        
        # Normalizar índice de mapeo
        def _norm_key(s):
            if s is None:
                return ''
            nk = unicodedata.normalize('NFKD', str(s))
            nk = nk.encode('ASCII', 'ignore').decode('ASCII')
            return nk.strip().upper()
        
        mapping_norm = {_norm_key(k): v for k, v in mapping.items()}
        
        # Extraer mapeo de grupos por color de fondo desde el archivo de horarios
        grupos_map = _extraer_grupos_por_color('Horario semana.xlsx')
        
        # Estilos reutilizables
        header_fill = PatternFill(start_color="1E4A80", end_color="1E4A80", fill_type="solid")
        gray_fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid") # Gris más flojo
        entry_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        group1_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
        group2_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        desc_fill = PatternFill(start_color="B8FF71", end_color="B8FF71", fill_type="solid")
        special_hour_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
        entry_half_fill = PatternFill(start_color="4EA72E", end_color="4EA72E", fill_type="solid")
        exit_half_fill = PatternFill(start_color="4EA72E", end_color="4EA72E", fill_type="solid")
        entry_half_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        special_hours = ['7-8', '8-9', '9-10', '21-22']
        header_font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
        emp_font = Font(name='Calibri', size=18)
        entry_font = Font(name='Calibri', size=16, bold=True)
        body_font = Font(name='Calibri', size=11)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_alignment = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)
        emp_alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
        thin_border = Border(left=Side(style='thin', color='808080'), 
                             right=Side(style='thin', color='808080'), 
                             top=Side(style='thin', color='808080'), 
                             bottom=Side(style='thin', color='808080'))
        emp_border = Border(left=Side(style='thin', color='808080'),
                            right=Side(style='thin', color='808080'),
                            top=Side(style='thin', color='9BC2E6'),
                            bottom=Side(style='thin', color='9BC2E6'))
        
        # Procesar cada día
        for dia, cuadrante_dict in cuadrantes_por_dia.items():
            empleados_dia = empleados_por_dia.get(dia)
            if empleados_dia is None:
                continue
            
            # CALCULAR DESCANSOS
            descansos_map = calcular_descansos(empleados_dia)
            
            # Generar DataFrame del cuadrante
            df = generar_cuadrante_buscador(empleados_dia, cuadrante_dict)
            
            # Crear hoja o limpiarla si ya existe
            ws = wb.create_sheet(title=dia)
            ws.sheet_view.zoomScale = 65
            ws.row_dimensions[1].height = 60
            
            # --- ENCABEZADOS (División Horaria Binaria: 2 columnas por hora) ---
            
            # 1. Empleado
            ws.cell(row=1, column=1, value='Empleado').font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=1).border = thin_border
            ws.cell(row=1, column=1).alignment = center_alignment
            
            # 2. HORA ENTR
            ws.cell(row=1, column=2, value='HORA ENTR').font = header_font
            ws.cell(row=1, column=2).fill = header_fill
            ws.cell(row=1, column=2).border = thin_border
            ws.cell(row=1, column=2).alignment = center_alignment
            
            hour_labels = [f"{h}-{h+1}" for h in range(7, 22)]
            
            # 3. Horas (2 columnas por hora)
            for i, h_label in enumerate(hour_labels):
                col_start = 3 + (i * 2)
                col_end = col_start + 1
                
                # Escribir en la primera y mergear
                c1 = ws.cell(row=1, column=col_start, value=h_label)
                c1.font = header_font
                c1.fill = header_fill
                c1.border = thin_border
                c1.alignment = header_alignment
                
                c2 = ws.cell(row=1, column=col_end)
                c2.fill = header_fill
                c2.border = thin_border
                
                ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
            
            # 4. Descanso
            col_desc = 3 + (len(hour_labels) * 2)
            ws.cell(row=1, column=col_desc, value='Desc').font = header_font
            ws.cell(row=1, column=col_desc).fill = header_fill
            ws.cell(row=1, column=col_desc).border = thin_border
            ws.cell(row=1, column=col_desc).alignment = center_alignment
            
            # Escribir datos con estilos
            for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
                # Empleado
                emp_cell = ws.cell(row=r_idx, column=1, value=row['Empleado'])
                emp_cell.font = emp_font
                emp_cell.alignment = emp_alignment
                emp_cell.border = emp_border

                # Aplicar color según color de fondo en Horario semana.xlsx
                emp_name_str = str(row['Empleado']).strip()
                grupo = grupos_map.get(emp_name_str, 0)
                if grupo == 1:
                    emp_cell.fill = group1_fill
                elif grupo == 2:
                    emp_cell.fill = group2_fill
                else:
                    emp_cell.fill = gray_fill

                ws.row_dimensions[r_idx].height = 23.4
                
                # HORA ENTR
                entr_cell = ws.cell(row=r_idx, column=2, value=row['HORA ENTR'])
                entr_cell.font = entry_font
                entr_cell.fill = entry_fill
                entr_cell.alignment = center_alignment
                entr_cell.number_format = 'HH:mm'
                entr_cell.border = thin_border

                # Parsear hora de entrada para lógica de media hora
                hora_entr_val = row.get('HORA ENTR')
                entry_h = -1
                entry_m = 0
                try:
                    s_entr = str(hora_entr_val).strip()
                    if ':' in s_entr:
                        parts = s_entr.split(':')
                        entry_h = int(parts[0])
                        entry_m = int(parts[1])
                except:
                    pass

                # Parsear hora de salida para lógica de media hora
                hora_salida_val = row.get('HORA SALIDA')
                exit_h = -1
                exit_m = 0
                try:
                    s_salida = str(hora_salida_val).strip()
                    if ':' in s_salida:
                        parts = s_salida.split(':')
                        exit_h = int(parts[0])
                        exit_m = int(parts[1])
                except:
                    pass

                # Horas (columnas 3 en adelante, 2 por hora)
                for i, h_label in enumerate(hour_labels):
                    col_start = 3 + (i * 2)
                    col_end = col_start + 1
                    
                    c1 = ws.cell(row=r_idx, column=col_start)
                    c2 = ws.cell(row=r_idx, column=col_end)
                    
                    c1.border = thin_border
                    c2.border = thin_border
                    
                    if h_label in special_hours:
                        c1.fill = special_hour_fill
                        c2.fill = special_hour_fill
                    
                    val = row.get(h_label, '')
                    
                    # Lógica para entrada a y 30: marcar segunda celda con X verde
                    current_h = 7 + i
                    if current_h == entry_h and entry_m == 30:
                        c1.value = ""
                        c2.value = "X"
                        c2.font = entry_half_font
                        c2.fill = entry_half_fill
                        c2.alignment = center_alignment
                        continue
                    
                    # Lógica para salida a y 30: marcar primera celda con X verde
                    if current_h == exit_h and exit_m == 30:
                        c1.value = "X"
                        c1.font = entry_half_font
                        c1.fill = exit_half_fill
                        c1.alignment = center_alignment
                        c2.value = ""
                        continue
                    
                    # Merge siempre para cubrir las dos celdas de 30 min
                    ws.merge_cells(start_row=r_idx, start_column=col_start, end_row=r_idx, end_column=col_end)
                    
                    if val is None or str(val).strip() == '' or str(val).strip().upper() == 'OFF':
                        continue
                    
                    key = str(val).strip()
                    style_entry = mapping.get(key) or mapping_norm.get(_norm_key(key))
                    
                    if style_entry:
                        c1.value = style_entry['nombre']
                        s = style_entry['style']
                        try:
                            f = copy(s['font'])
                            if f: f.name = 'Calibri'
                            c1.font = f
                            
                            fill_copy = copy(s['fill'])
                            c1.fill = fill_copy
                            c2.fill = fill_copy
                            
                            border_copy = copy(s['border']) if s['border'] and s['border'].left.style else thin_border
                            c1.border = border_copy
                            c2.border = border_copy
                            
                            c1.alignment = center_alignment
                            c1.number_format = s.get('number_format', c1.number_format)
                        except Exception:
                            c1.alignment = center_alignment
                    else:
                        c1.value = val
                        c1.font = body_font
                        c1.alignment = center_alignment

                # Columna de Descanso
                descanso_val = descansos_map.get(row['Empleado'], '')
                desc_cell = ws.cell(row=r_idx, column=col_desc, value=descanso_val)
                desc_cell.font = body_font
                desc_cell.fill = desc_fill
                desc_cell.alignment = center_alignment
                desc_cell.border = thin_border
            
            # Ancho de columnas
            ws.column_dimensions['A'].width = 48
            ws.column_dimensions['B'].width = 12
            for c in range(3, col_desc): # Horas (ahora el doble de columnas)
                ws.column_dimensions[get_column_letter(c)].width = 6 # Mitad de 12
            ws.column_dimensions[get_column_letter(col_desc)].width = 10 # Descanso
            
            print(f"  [OK] Pestaña '{dia}' generada ({len(df)} empleados)")
        
        # Guardar en el buffer o ruta indicada
        wb.save(ruta_salida)
        print(f"\n[OK] Salida semanal generada exitosamente")
        return True
    
    except Exception as e:
        print(f"[ERROR] No se pudo generar la salida: {e}")
        return False


def generar_salida(empleados_hoy, cuadrante_dict, ruta_salida='Zonning_Hoy.xlsx'):
    """
    Función principal: genera DataFrame y lo escribe a Excel.
    
    Retorna:
        bool: True si exitoso
    """
    # Generar DataFrame del cuadrante
    df = generar_cuadrante_buscador(empleados_hoy, cuadrante_dict)
    
    # Escribir a Excel
    return escribir_excel(df, ruta_salida)


if __name__ == '__main__':
    # Prueba rápida (requiere lector_datos.py y motor_logica.py)
    from lector_datos import cargar_todo
    from motor_logica import asignar
    
    datos = cargar_todo()
    if datos:
        empleados_por_dia = datos['empleados_por_dia']
        zonas = datos['zonas']
        mapeo_hab = datos['mapeo_habilidades']
        
        cuadrantes_por_dia = {}
        for dia, empleados_dia in empleados_por_dia.items():
            cuadrante = asignar(empleados_dia, zonas, mapeo_hab)
            cuadrantes_por_dia[dia] = cuadrante
        
        exito = generar_salida_semanal(empleados_por_dia, cuadrantes_por_dia)
        if exito:
            print(f"[OK] Salida semanal generada correctamente")
